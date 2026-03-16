import os
from dataclasses import dataclass
from glob import glob

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side


@dataclass
class FormatExcel:
    """Class to format Excel files by adjusting cell alignment, borders, and column widths.

    Attributes:
        path (str): The directory or file path of the Excel file(s) to format.
    """

    path: str

    def load_workbook(self, file_path: str):
        """Loads an Excel workbook from the specified file path.

        Args:
            file_path (str): The path to the Excel file to load.
        """
        self.wb = load_workbook(file_path)
        self.file_path = file_path

    def align(self):
        """Aligns the text in the Excel cells.

        - Aligns text in the first column to the left.
        - Aligns text in all other columns to the center.
        """
        for ws in self.wb.worksheets:
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

    def line(self):
        """Applies borders to the cells in the Excel workbook.

        - Creates thin borders for the top of the first row and bottom of the last row.
        - Creates a thick border for the bottom of the first row.
        """
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="thick", color="000000")  # noqa: F841

        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border()

            max_row = ws.max_row

            for cell in ws[1]:
                cell.border = Border(top=thin, bottom=thin)

            for cell in ws[max_row]:
                cell.border = Border(bottom=thin)

    def convert_brackets(self, direction="to_square"):
        """Converts parentheses () to square brackets [] and vice versa in the Excel cells.

        This method goes through all the cells in the workbook and:
        - Converts () to [] if direction is 'to_square'.
        - Converts [] to () if direction is 'to_round'.

        Args:
            direction (str): Direction of the bracket conversion.
                            - 'to_square' for converting () to [].
                            - 'to_round' for converting [] to ().
                            Default is 'to_square'.
        """
        # Validate the direction parameter
        if direction not in ["to_square", "to_round"]:
            raise ValueError(
                "Invalid direction. Choose either 'to_square' or 'to_round'."
            )

        for ws in self.wb.worksheets:
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    if isinstance(
                        cell.value, str
                    ):  # Check if the cell contains a string
                        if direction == "to_square":
                            cell.value = cell.value.replace("(", "[").replace(")", "]")
                        elif direction == "to_round":
                            cell.value = cell.value.replace("[", "(").replace("]", ")")

    def auto_adjust_column_widths(self):
        """Automatically adjusts the column widths to fit the contents."""
        for ws in self.wb.worksheets:
            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter  # Get the column letter
                for cell in column:
                    try:
                        cell_length = sum(
                            2 if ord(char) > 127 else 1 for char in str(cell.value)
                        )
                        max_length = max(max_length, cell_length)
                    except Exception:
                        pass
                # Adjust for header row
                # header_cell = ws[f"{col_letter}1"]
                # if header_cell.value:
                #     max_length = max(max_length, len(str(header_cell.value)))
                adjusted_width = max_length + 2  # Add some padding to the width
                ws.column_dimensions[col_letter].width = adjusted_width

    def save(self):
        """Saves the currently loaded workbook to its file path."""
        self.wb.save(self.file_path)

    def process(
        self, align=True, line=True, convert_brackets=False, adjust_col_widths=False
    ):
        """Processes and formats Excel files.

        - If the provided path is a directory, it formats all Excel files in that directory.
        - If the provided path is a file, it formats that specific Excel file.

        Args:
            align (bool): Whether to apply text alignment. Default is True.
            line (bool): Whether to apply borders. Default is True.
            convert_brackets (bool): Whether to convert brackets. Default is False.
            adjust_col_widths (bool): Whether to adjust column widths. Default is False.
        """
        files_to_process = (
            glob(os.path.join(self.path, "*.xlsx"))
            if os.path.isdir(self.path)
            else [self.path]
        )

        for file in files_to_process:
            self.load_workbook(file)
            if align:
                self.align()
            if line:
                self.line()
            if convert_brackets:
                self.convert_brackets()
            if adjust_col_widths:
                self.auto_adjust_column_widths()
            self.save()


if __name__ == "__main__":
    path = "..."
    excel_formatter = FormatExcel(path)
    excel_formatter.process()
